import customtkinter
import tkinter as tk
import openpyxl as xl
from tkinter import *
from tkinter import messagebox 
from datetime import *

app = customtkinter.CTk()
app.title('Supermrket')
app.geometry('800x500')
app.config(bg='#121111')


products={
    'milk':3,
    'soda':2,
    'shampoo':5,
    'soap':4,
    'Chips':1.23,
    'eggs':7,
    "Apple": 1.99,
    "Banana": 0.99,
    "Guava": 1.49,
    "Strawberry": 2.49,
    "Kiwi": 2.99,
    "Orange": 1.49,
    'meat':12,
    'fish':8,
    'prawns':5,
    'lambmeat':9,
    'chicken':10.23,
    'mince':7 
}

font1=('Arial',32,'bold')
font2=('Helvetica',18,'bold')
font3=('Helvetica',25,'bold')

variable1 = None  
variable2 = None  
variable3 = None  
variable4 = None  

def new_customer():
    variable1.set('0')
    variable2.set('0')
    variable3.set('0')
    variable4.set('0')
    customer_entry.delete(0, tk.END)  # Use tk.END instead of just END
    bill_result_label.configure(text='')
    date_result_label.configure(text='')
    time_result_label.configure(text='')
    name_result_label.configure(text='')
    total_result_label.configure(text='')

def checkout():
    total_price = 0  # Initialize the total price variable

    # Calculate the total price based on the quantity and price of each item
    total_price += int(variable1.get()) * products['milk']
    total_price += int(variable1.get()) * products['soda']
    total_price += int(variable1.get()) * products['shampoo']
    total_price += int(variable1.get()) * products['soap']
    total_price += int(variable1.get()) * products['Chips']
    total_price += int(variable1.get()) * products['eggs']
    total_price += int(variable1.get()) * products['Apple']
    total_price += int(variable1.get()) * products['Banana']
    total_price += int(variable1.get()) * products['Orange']
    total_price += int(variable1.get()) * products['Guava']
    total_price += int(variable1.get()) * products['Kiwi']
    total_price += int(variable1.get()) * products['Strawberry']
    total_price += int(variable1.get()) * products['meat']
    total_price += int(variable1.get()) * products['lambmeat']
    total_price += int(variable1.get()) * products['chicken']
    total_price += int(variable1.get()) * products['mince']
    total_price += int(variable1.get()) * products['prawns']
    total_price += int(variable1.get()) * products['fish']    
    if total_price !=0 and customer_entry.get()!='':
        file=xl.load_workbook("mydata.xlsx")
        sheet=file['Sheet1']
        bill_result_label.configure(text=sheet.max_row)
        date_result_label.configure(text=date.today())
        time_result_label.configure(text=datetime.now().time().strftime("%I:%M:%S %p"))
        name_result_label.configure(text=customer_entry.get())
        total_result_label.configure(text=f'{total_price}$')
        sheet.cell(column=1,row=sheet.max_row+1,value=sheet.max_row)
        sheet.cell(column=2,row=sheet.max_row,value=date.today())
        sheet.cell(column=3,row=sheet.max_row,value=datetime.now().time())
        sheet.cell(column=4,row=sheet.max_row,value=customer_entry.get())
        sheet.cell(column=5,row=sheet.max_row,value=total_price)
        file.save('mydata.xlsx')
        messagebox.showinfo('Sucess','Data has been Saved.')
    else:
        messagebox.showerror('Error','Enter all the data')

# Define global variables to store the PhotoImage objects
image3 = None
image4 = None
image5 = None
image6 = None
image7 = None
image8 = None
image9 = None
image10 = None
image11 = None
image12 = None
image13 = None
image14 = None
image15 = None
image16 = None
image17 = None
image18 = None
image19 = None
image20 = None
variable1 = None  
variable2 = None  
variable3 = None  
variable4 = None 


def display_grocery():
    global image3, image4, image5, image6, image7, image8, variable1, variable2, variable3, variable4

    p1_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p1_frame.place(x=10,y=120)

    image3=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\download1.png")
    image3_label=Label(p1_frame,image=image3,bg='#252625',width=200,height=200)
    image3_label.place(x=10,y=10)

    p1_name_label=customtkinter.CTkLabel(p1_frame,font=font2,text='milk',text_color='#fff',bg_color='#252625')
    p1_name_label.place(x=250,y=100)

    p1_price_label=customtkinter.CTkLabel(p1_frame,font=font2,text=f'Price:{products["milk"]}$',text_color='#fff',bg_color='#252625')
    p1_price_label.place(x=250,y=120)

    p1_quantity_label=customtkinter.CTkLabel(p1_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p1_quantity_label.place(x=250,y=140)
 
    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()




    p1_quantity_option=customtkinter.CTkComboBox(p1_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p1_quantity_option.set('0')
    p1_quantity_option.place(x=250,y=170)


    p2_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p2_frame.place(x=400,y=120)

    image4=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\soda.png")
    image4_label=Label(p2_frame,image=image4,bg='#252625',width=200,height=200)
    image4_label.place(x=10,y=10)

    p2_name_label=customtkinter.CTkLabel(p2_frame,font=font2,text='soda',text_color='#fff',bg_color='#252625')
    p2_name_label.place(x=250,y=100)

    p2_price_label=customtkinter.CTkLabel(p2_frame,font=font2,text=f'Price:{products["soda"]}$',text_color='#fff',bg_color='#252625')
    p2_price_label.place(x=250,y=120)

    p2_quantity_label=customtkinter.CTkLabel(p2_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p2_quantity_label.place(x=250,y=140)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()
    

    p2_quantity_option=customtkinter.CTkComboBox(p2_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p2_quantity_option.set('0')
    p2_quantity_option.place(x=250,y=170)

    p3_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p3_frame.place(x=10,y=400)

    image5=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\shampoo.png")
    image5_label=Label(p3_frame,image=image5,bg='#252625',width=200,height=200)
    image5_label.place(x=10,y=10)

    p3_name_label=customtkinter.CTkLabel(p3_frame,font=font2,text='shampoo',text_color='#fff',bg_color='#252625')
    p3_name_label.place(x=250,y=100)

    p3_price_label=customtkinter.CTkLabel(p3_frame,font=font2,text=f'Price:{products["shampoo"]}$',text_color='#fff',bg_color='#252625')
    p3_price_label.place(x=250,y=130)

    p3_quantity_label=customtkinter.CTkLabel(p3_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p3_quantity_label.place(x=250,y=150)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()


    p3_quantity_option=customtkinter.CTkComboBox(p3_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p3_quantity_option.set('0')
    p3_quantity_option.place(x=250,y=180)


    p4_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p4_frame.place(x=400,y=400)

    image6=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\soap.png")
    image6_label=Label(p4_frame,image=image6,bg='#252625',width=200,height=200)
    image6_label.place(x=10,y=10)

    p4_name_label=customtkinter.CTkLabel(p4_frame,font=font2,text='soap',text_color='#fff',bg_color='#252625')
    p4_name_label.place(x=250,y=100)

    p4_price_label=customtkinter.CTkLabel(p4_frame,font=font2,text=f'Price:{products["soap"]}$',text_color='#fff',bg_color='#252625')
    p4_price_label.place(x=250,y=130)
 
    p4_quantity_label=customtkinter.CTkLabel(p4_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p4_quantity_label.place(x=250,y=150)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()


    p4_quantity_option=customtkinter.CTkComboBox(p4_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p4_quantity_option.set('0')
    p4_quantity_option.place(x=250,y=180)

    p5_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p5_frame.place(x=800,y=120)

    image7=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\chips.png")
    image7_label=Label(p5_frame,image=image7,bg='#252625',width=200,height=200)
    image7_label.place(x=10,y=10)

    p5_name_label=customtkinter.CTkLabel(p5_frame,font=font2,text='chips',text_color='#fff',bg_color='#252625')
    p5_name_label.place(x=250,y=100)

    p5_price_label=customtkinter.CTkLabel(p5_frame,font=font2,text=f'Price:{products["Chips"]}$',text_color='#fff',bg_color='#252625')
    p5_price_label.place(x=250,y=120)

    p5_quantity_label=customtkinter.CTkLabel(p5_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p5_quantity_label.place(x=250,y=140)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()

    p5_quantity_option=customtkinter.CTkComboBox(p5_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p5_quantity_option.set('0')
    p5_quantity_option.place(x=250,y=170)
    
    p6_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p6_frame.place(x=800,y=400)

    image8=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\eggs.png")
    image8_label=Label(p6_frame,image=image8,bg='#252625',width=200,height=200)
    image8_label.place(x=10,y=10)

    p6_name_label=customtkinter.CTkLabel(p6_frame,font=font2,text='eggs',text_color='#fff',bg_color='#252625')
    p6_name_label.place(x=250,y=100)

    p6_price_label=customtkinter.CTkLabel(p6_frame,font=font2,text=f'Price:{products["eggs"]}$',text_color='#fff',bg_color='#252625')
    p6_price_label.place(x=250,y=130)

    p6_quantity_label=customtkinter.CTkLabel(p6_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p6_quantity_label.place(x=250,y=150)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()

    p6_quantity_option=customtkinter.CTkComboBox(p6_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p6_quantity_option.set('0')
    p6_quantity_option.place(x=250,y=180)
    
    
def display_fruits():
    global image9, image10, image11, image12, image13, image14,variable1, variable2, variable3, variable4
    p1_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p1_frame.place(x=10,y=120)

    image9=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\apple.png")
    image9_label=Label(p1_frame,image=image9,bg='#252625',width=200,height=200)
    image9_label.place(x=10,y=10)

    p1_name_label=customtkinter.CTkLabel(p1_frame,font=font2,text='apple',text_color='#fff',bg_color='#252625')
    p1_name_label.place(x=250,y=100)

    p1_price_label=customtkinter.CTkLabel(p1_frame,font=font2,text=f'Price:{products["Apple"]}$',text_color='#fff',bg_color='#252625')
    p1_price_label.place(x=250,y=120)

    p1_quantity_label=customtkinter.CTkLabel(p1_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p1_quantity_label.place(x=250,y=140)
 
    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()

    p1_quantity_option=customtkinter.CTkComboBox(p1_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p1_quantity_option.set('0')
    p1_quantity_option.place(x=250,y=170)


    p2_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p2_frame.place(x=400,y=120)

    image10=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\banana.png")
    image10_label=Label(p2_frame,image=image10,bg='#252625',width=220,height=220)
    image10_label.place(x=10,y=10)

    p2_name_label=customtkinter.CTkLabel(p2_frame,font=font2,text='banana',text_color='#fff',bg_color='#252625')
    p2_name_label.place(x=250,y=100)

    p2_price_label=customtkinter.CTkLabel(p2_frame,font=font2,text=f'Price:{products["Banana"]}$',text_color='#fff',bg_color='#252625')
    p2_price_label.place(x=250,y=120)

    p2_quantity_label=customtkinter.CTkLabel(p2_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p2_quantity_label.place(x=250,y=140)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()

    p2_quantity_option=customtkinter.CTkComboBox(p2_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p2_quantity_option.set('0')
    p2_quantity_option.place(x=250,y=170)

    p3_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p3_frame.place(x=10,y=400)

    image11=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\guava.png")
    image11_label=Label(p3_frame,image=image11,bg='#252625',width=200,height=200)
    image11_label.place(x=10,y=10)

    p3_name_label=customtkinter.CTkLabel(p3_frame,font=font2,text='guava',text_color='#fff',bg_color='#252625')
    p3_name_label.place(x=250,y=100)

    p3_price_label=customtkinter.CTkLabel(p3_frame,font=font2,text=f'Price:{products["Guava"]}$',text_color='#fff',bg_color='#252625')
    p3_price_label.place(x=250,y=130)

    p3_quantity_label=customtkinter.CTkLabel(p3_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p3_quantity_label.place(x=250,y=150)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()

    p3_quantity_option=customtkinter.CTkComboBox(p3_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p3_quantity_option.set('0')
    p3_quantity_option.place(x=250,y=180)


    p4_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p4_frame.place(x=400,y=400)

    image12=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\strawberry.png")
    image12_label=Label(p4_frame,image=image12,bg='#252625',width=200,height=200)
    image12_label.place(x=10,y=10)

    p4_name_label=customtkinter.CTkLabel(p4_frame,font=font2,text='strawberry',text_color='#fff',bg_color='#252625')
    p4_name_label.place(x=250,y=100)

    p4_price_label=customtkinter.CTkLabel(p4_frame,font=font2,text=f'Price:{products["Strawberry"]}$',text_color='#fff',bg_color='#252625')
    p4_price_label.place(x=250,y=130)
 
    p4_quantity_label=customtkinter.CTkLabel(p4_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p4_quantity_label.place(x=250,y=150)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()


    p4_quantity_option=customtkinter.CTkComboBox(p4_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p4_quantity_option.set('0')
    p4_quantity_option.place(x=250,y=180)

    p5_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p5_frame.place(x=800,y=120)

    image13=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\kiwi1.png")
    image13_label=Label(p5_frame,image=image13,bg='#252625',width=200,height=200)
    image13_label.place(x=10,y=10)

    p5_name_label=customtkinter.CTkLabel(p5_frame,font=font2,text='kiwi',text_color='#fff',bg_color='#252625')
    p5_name_label.place(x=250,y=100)

    p5_price_label=customtkinter.CTkLabel(p5_frame,font=font2,text=f'Price:{products["Kiwi"]}$',text_color='#fff',bg_color='#252625')
    p5_price_label.place(x=250,y=120)

    p5_quantity_label=customtkinter.CTkLabel(p5_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p5_quantity_label.place(x=250,y=140)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()

    p5_quantity_option=customtkinter.CTkComboBox(p5_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p5_quantity_option.set('0')
    p5_quantity_option.place(x=250,y=170)
    
    p6_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p6_frame.place(x=800,y=400)

    image14=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\orange.png")
    image14_label=Label(p6_frame,image=image14,bg='#252625',width=200,height=200)
    image14_label.place(x=10,y=10)

    p6_name_label=customtkinter.CTkLabel(p6_frame,font=font2,text='orange',text_color='#fff',bg_color='#252625')
    p6_name_label.place(x=250,y=100)

    p6_price_label=customtkinter.CTkLabel(p6_frame,font=font2,text=f'Price:{products["Orange"]}$',text_color='#fff',bg_color='#252625')
    p6_price_label.place(x=250,y=130)

    p6_quantity_label=customtkinter.CTkLabel(p6_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p6_quantity_label.place(x=250,y=150)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()

    p6_quantity_option=customtkinter.CTkComboBox(p6_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p6_quantity_option.set('0')
    p6_quantity_option.place(x=250,y=180)
   

def display_meat_items():
    global image15, image16, image17, image18, image19, image20, variable1, variable2, variable3, variable4

    p1_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p1_frame.place(x=10,y=120)

    image15=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\beef.png")
    image15_label=Label(p1_frame,image=image15,bg='#252625',width=230,height=200)
    image15_label.place(x=10,y=10)

    p1_name_label=customtkinter.CTkLabel(p1_frame,font=font2,text='Beef Meat',text_color='#fff',bg_color='#252625')
    p1_name_label.place(x=250,y=100)

    p1_price_label=customtkinter.CTkLabel(p1_frame,font=font2,text=f'Price:{products["meat"]}$',text_color='#fff',bg_color='#252625')
    p1_price_label.place(x=250,y=120)

    p1_quantity_label=customtkinter.CTkLabel(p1_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p1_quantity_label.place(x=250,y=140)
 
    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()




    p1_quantity_option=customtkinter.CTkComboBox(p1_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p1_quantity_option.set('0')
    p1_quantity_option.place(x=250,y=170)


    p2_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p2_frame.place(x=400,y=120)

    image16=PhotoImage(file="C:\\Users\\dell\\\Desktop\\labs\\chicken.png")
    image16_label=Label(p2_frame,image=image16,bg='#252625',width=230,height=200)
    image16_label.place(x=10,y=10)

    p2_name_label=customtkinter.CTkLabel(p2_frame,font=font2,text='chicken',text_color='#fff',bg_color='#252625')
    p2_name_label.place(x=250,y=100)

    p2_price_label=customtkinter.CTkLabel(p2_frame,font=font2,text=f'Price:{products["chicken"]}$',text_color='#fff',bg_color='#252625')
    p2_price_label.place(x=250,y=120)

    p2_quantity_label=customtkinter.CTkLabel(p2_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p2_quantity_label.place(x=250,y=140)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()
    

    p2_quantity_option=customtkinter.CTkComboBox(p2_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p2_quantity_option.set('0')
    p2_quantity_option.place(x=250,y=170)

    p3_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p3_frame.place(x=10,y=400)

    image17=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\mince.png")
    image17_label=Label(p3_frame,image=image17,bg='#252625',width=200,height=200)
    image17_label.place(x=10,y=10)

    p3_name_label=customtkinter.CTkLabel(p3_frame,font=font2,text='mince',text_color='#fff',bg_color='#252625')
    p3_name_label.place(x=250,y=100)

    p3_price_label=customtkinter.CTkLabel(p3_frame,font=font2,text=f'Price:{products["mince"]}$',text_color='#fff',bg_color='#252625')
    p3_price_label.place(x=250,y=130)

    p3_quantity_label=customtkinter.CTkLabel(p3_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p3_quantity_label.place(x=250,y=150)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()


    p3_quantity_option=customtkinter.CTkComboBox(p3_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p3_quantity_option.set('0')
    p3_quantity_option.place(x=250,y=180)


    p4_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p4_frame.place(x=400,y=400)

    image18=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\fish1.png")
    image18_label=Label(p4_frame,image=image18,bg='#252625',width=230,height=200)
    image18_label.place(x=10,y=10)

    p4_name_label=customtkinter.CTkLabel(p4_frame,font=font2,text='fish',text_color='#fff',bg_color='#252625')
    p4_name_label.place(x=250,y=100)

    p4_price_label=customtkinter.CTkLabel(p4_frame,font=font2,text=f'Price:{products["fish"]}$',text_color='#fff',bg_color='#252625')
    p4_price_label.place(x=250,y=130)
 
    p4_quantity_label=customtkinter.CTkLabel(p4_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p4_quantity_label.place(x=250,y=150)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()


    p4_quantity_option=customtkinter.CTkComboBox(p4_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p4_quantity_option.set('0')
    p4_quantity_option.place(x=250,y=180)

    p5_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p5_frame.place(x=800,y=120)

    image19=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\lamb meat.png")
    image19_label=Label(p5_frame,image=image19,bg='#252625',width=200,height=200)
    image19_label.place(x=10,y=10)

    p5_name_label=customtkinter.CTkLabel(p5_frame,font=font2,text='lamb meat',text_color='#fff',bg_color='#252625')
    p5_name_label.place(x=250,y=100)

    p5_price_label=customtkinter.CTkLabel(p5_frame,font=font2,text=f'Price:{products["lambmeat"]}$',text_color='#fff',bg_color='#252625')
    p5_price_label.place(x=250,y=120)

    p5_quantity_label=customtkinter.CTkLabel(p5_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p5_quantity_label.place(x=250,y=140)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()

    p5_quantity_option=customtkinter.CTkComboBox(p5_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p5_quantity_option.set('0')
    p5_quantity_option.place(x=250,y=170)
    
    p6_frame= customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252625',corner_radius=30,height=250,width=360)
    p6_frame.place(x=800,y=400)

    image20=PhotoImage(file="C:\\Users\\dell\\Desktop\\labs\\prawns.png")
    image20_label=Label(p6_frame,image=image20,bg='#252625',width=200,height=200)
    image20_label.place(x=10,y=10)

    p6_name_label=customtkinter.CTkLabel(p6_frame,font=font2,text='prawns',text_color='#fff',bg_color='#252625')
    p6_name_label.place(x=250,y=100)

    p6_price_label=customtkinter.CTkLabel(p6_frame,font=font2,text=f'Price:{products["prawns"]}$',text_color='#fff',bg_color='#252625')
    p6_price_label.place(x=250,y=130)

    p6_quantity_label=customtkinter.CTkLabel(p6_frame,font=font2,text='Quantity:',text_color='#fff',bg_color='#252625')
    p6_quantity_label.place(x=250,y=150)

    quantity_option=['0','1','2','3']
    variable1 = StringVar()
    variable2 = StringVar()
    variable3 = StringVar()
    variable4 = StringVar()

    p6_quantity_option=customtkinter.CTkComboBox(p6_frame,font=font2,text_color='#06911f',values=quantity_option,variable=variable1,state='readonly',button_color='#f75d05',width=60)
    p6_quantity_option.set('0')
    p6_quantity_option.place(x=250,y=180)
    
    
# Create the menu
menu_frame = tk.Frame(app, bg='black')
menu_frame.pack()

# Create the menu buttons
grocery_btn = tk.Button(menu_frame, text="Grocery",command=display_grocery, font=font2, fg='#fff', bg='#121111', cursor='hand2', width=20)
grocery_btn.pack(side=tk.LEFT)

fruits_btn = tk.Button(menu_frame, text="Fruits",command=display_fruits, font=font2, fg='#fff', bg='#121111', cursor='hand2', width=20)
fruits_btn.pack(side=tk.LEFT)



meat_items_btn = tk.Button(menu_frame, text="Meat Items",command=display_meat_items ,font=font2, fg='#fff', bg='#121111', cursor='hand2', width=20)
meat_items_btn.pack(side=tk.LEFT)


 
customer_label=customtkinter.CTkLabel(app,font=font3,text='Customer Name:',text_color='#fff',bg_color='#121111')
customer_label.place(x=50,y=750)

customer_entry=customtkinter.CTkEntry(app,font=font2,text_color='#000',fg_color='#fff',bg_color='#121111',border_color='#fff',width=220)
customer_entry.place(x=270,y=750)

checkout_button=customtkinter.CTkButton(app,command=checkout,font=font2,text_color='#fff',text='Checkout',fg_color='#02a81d',hover_color='#018517',bg_color='#121111',cursor='hand2',corner_radius=8,width=150)
checkout_button.place(x=80,y=800)

new_customer_button=customtkinter.CTkButton(app,command=new_customer,font=font2,text_color='#fff',text='New Customer',fg_color='#bd02b0',hover_color='#94038a',bg_color='#121111',cursor='hand2',corner_radius=8,width=150)
new_customer_button.place(x=250,y=800)

bill_frame=customtkinter.CTkFrame(app,bg_color='#121111',fg_color='#252652',border_color='#fff',border_width=3,width=300,height=250)
bill_frame.place(x=950,y=700)

bill_number_label=customtkinter.CTkLabel(bill_frame,font=font2,text='Bill Number: ',text_color='#fff',bg_color='#252625')
bill_number_label.place(x=10,y=10)

date_label=customtkinter.CTkLabel(bill_frame,font=font2,text='Date: ',text_color='#fff',bg_color='#252625')
date_label.place(x=10,y=50)

time_label=customtkinter.CTkLabel(bill_frame,font=font2,text='Time: ',text_color='#fff',bg_color='#252625')
time_label.place(x=10,y=90)

name_label=customtkinter.CTkLabel(bill_frame,font=font2,text='Customer Name: ',text_color='#fff',bg_color='#252625')
name_label.place(x=10,y=130)

total_label=customtkinter.CTkLabel(bill_frame,font=font2,text='Total: ',text_color='#fff',bg_color='#252625')
total_label.place(x=10,y=170)

bill_result_label=customtkinter.CTkLabel(bill_frame,font=font2,text='',text_color='#fff',bg_color='#252625')
bill_result_label.place(x=180,y=10)

date_result_label=customtkinter.CTkLabel(bill_frame,font=font2,text='',text_color='#fff',bg_color='#252625')
date_result_label.place(x=180,y=50)

time_result_label=customtkinter.CTkLabel(bill_frame,font=font2,text='',text_color='#fff',bg_color='#252625')
time_result_label.place(x=180,y=90)

name_result_label=customtkinter.CTkLabel(bill_frame,font=font2,text='',text_color='#fff',bg_color='#252625')
name_result_label.place(x=180,y=130)

total_result_label=customtkinter.CTkLabel(bill_frame,font=font2,text='',text_color='#fff',bg_color='#252625')
total_result_label.place(x=180,y=170)





app.mainloop()
